from __future__ import annotations

import io
import json

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def _project(client: TestClient) -> str:
    response = client.post("/api/projects", json={"name": "Acme Security Review"})
    assert response.status_code == 201
    return response.json()["id"]


def test_workspace_and_health_expose_the_supported_demo_contract(client: TestClient) -> None:
    health = client.get("/api/health")
    workspace = client.get("/")

    assert health.status_code == 200
    assert health.json()["provider"]["active"] == "deterministic-demo"
    assert workspace.status_code == 200
    assert 'accept=".xlsx,.csv,.pdf,.docx,.txt,.md"' in workspace.text
    assert 'accept=".xlsx,.xls,' not in workspace.text
    assert "evidence coverage" in workspace.text
    assert "Model provider connected" not in workspace.text


def test_synthetic_demo_endpoint_seeds_real_project_files(client: TestClient) -> None:
    seeded = client.post("/api/demo")

    assert seeded.status_code == 201
    payload = seeded.json()
    assert payload["project"]["name"] == "CloudDesk Synthetic Review"
    assert len(payload["questions"]) == 8
    assert {document["kind"] for document in payload["documents"]} == {"questionnaire", "evidence"}

    run = client.post(f"/api/projects/{payload['project']['id']}/run")
    assert run.status_code == 200
    result = run.json()
    assert result["processed"] == 8
    assert result["draft_count"] == 7
    assert result["needs_evidence_count"] == 1
    conflict = next(question for question in result["questions"] if "incident notification" in question["question"])
    missing = next(question for question in result["questions"] if "subprocessors" in question["question"])
    assert conflict["checks"]["contradictions"]
    assert missing["answer"] is None


def test_end_to_end_grounded_draft_review_and_export(client: TestClient) -> None:
    project_id = _project(client)
    upload = client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("security.md", b"MFA is enabled for all administrator accounts.\n", "text/markdown"))],
    )
    assert upload.status_code == 201
    added = client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["Is MFA enabled for administrator accounts?"]},
    )
    question_id = added.json()["questions"][0]["id"]

    run = client.post(f"/api/projects/{project_id}/run")
    payload = run.json()
    assert payload["draft_count"] == 1
    question = payload["questions"][0]
    assert question["checks"]["grounded"] is True
    assert question["citations"][0]["document"] == "security.md"
    assert question["citations"][0]["quote"] == "MFA is enabled for all administrator accounts."

    review = client.patch(
        f"/api/projects/{project_id}/questions/{question_id}/review",
        json={"action": "approve", "note": "Confirmed by security lead"},
    )
    assert review.status_code == 200
    assert review.json()["status"] == "approved"

    exported = client.get(f"/api/projects/{project_id}/export?format=json")
    rows = json.loads(exported.content)
    assert rows[0]["status"] == "approved"
    assert rows[0]["citations"][0]["document"] == "security.md"

    workbook_response = client.get(f"/api/projects/{project_id}/export?format=xlsx")
    workbook = load_workbook(io.BytesIO(workbook_response.content))
    assert workbook.active["A2"].value == "Is MFA enabled for administrator accounts?"


def test_export_excludes_unapproved_answers_by_default(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("security.txt", b"MFA is enabled for administrators.\n", "text/plain"))],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["Is MFA enabled for administrators?"]},
    )
    client.post(f"/api/projects/{project_id}/run")

    assert json.loads(client.get(f"/api/projects/{project_id}/export?format=json").content) == []
    drafts = json.loads(
        client.get(f"/api/projects/{project_id}/export?format=json&include_drafts=true").content
    )
    assert len(drafts) == 1


def test_rejected_answer_can_return_to_draft_and_be_regenerated(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("security.txt", b"MFA is enabled for administrators.\n", "text/plain"))],
    )
    added = client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["Is MFA enabled for administrators?"]},
    ).json()
    question_id = added["questions"][0]["id"]
    client.post(f"/api/projects/{project_id}/run")

    rejected = client.patch(
        f"/api/projects/{project_id}/questions/{question_id}/review",
        json={"action": "reject", "note": "Tighten the wording."},
    )
    assert rejected.json()["status"] == "rejected"

    reopened = client.patch(
        f"/api/projects/{project_id}/questions/{question_id}",
        json={"edited_answer": "MFA is enabled for administrators.", "note": "Revised."},
    )
    assert reopened.json()["status"] == "draft"

    client.patch(
        f"/api/projects/{project_id}/questions/{question_id}/review",
        json={"action": "reject", "note": "Regenerate from evidence."},
    )
    rerun = client.post(f"/api/projects/{project_id}/run").json()
    assert rerun["processed"] == 1
    assert rerun["questions"][0]["status"] == "draft"


def test_approval_rejects_unsupported_human_edit(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("security.txt", b"MFA is enabled for administrators.\n", "text/plain"))],
    )
    added = client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["Is MFA enabled for administrators?"]},
    ).json()
    question_id = added["questions"][0]["id"]
    client.post(f"/api/projects/{project_id}/run")

    response = client.patch(
        f"/api/projects/{project_id}/questions/{question_id}/review",
        json={"action": "approve", "edited_answer": "MFA is enabled and audited every 7 days."},
    )

    assert response.status_code == 422
    assert "not supported" in response.json()["detail"]

    polarity_reversal = client.patch(
        f"/api/projects/{project_id}/questions/{question_id}/review",
        json={"action": "approve", "edited_answer": "MFA is not enabled for administrators."},
    )
    assert polarity_reversal.status_code == 422


def test_missing_evidence_does_not_invent_an_answer(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("privacy.txt", b"The privacy contact is documented internally.\n", "text/plain"))],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["What is the disaster recovery RTO?"]},
    )

    question = client.post(f"/api/projects/{project_id}/run").json()["questions"][0]

    assert question["status"] == "needs_evidence"
    assert question["answer"] is None
    missing = client.get(f"/api/projects/{project_id}/missing-evidence").json()["items"]
    assert len(missing) == 1


def test_generic_production_terms_do_not_mask_missing_subprocessor_evidence(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[
            (
                "files",
                (
                    "security.txt",
                    b"Production administrative access requires multi-factor authentication.\n",
                    "text/plain",
                ),
            )
        ],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["Provide the current list of production subprocessors."]},
    )

    question = client.post(f"/api/projects/{project_id}/run").json()["questions"][0]

    assert question["status"] == "needs_evidence"
    assert question["answer"] is None


def test_verify_evidence_tool_checks_stored_quote_integrity(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("policy.txt", b"Backups are retained for 30 days.\n", "text/plain"))],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["How long are backups retained?"]},
    )
    question = client.post(f"/api/projects/{project_id}/run").json()["questions"][0]
    citation = question["citations"][0]

    verified = client.post(
        "/api/tools/verify-evidence",
        json={
            "tenant_id": "demo",
            "project_id": project_id,
            "question": question["question"],
            "answer": question["answer"],
            "citations": [citation],
        },
    )
    assert verified.status_code == 200
    assert verified.json()["grounded"] is True
    assert verified.json()["citation_integrity"][0]["verified"] is True

    fabricated = dict(citation, quote="Backups are retained for 365 days.")
    rejected = client.post(
        "/api/tools/verify-evidence",
        json={
            "tenant_id": "demo",
            "project_id": project_id,
            "question": question["question"],
            "answer": "Backups are retained for 365 days.",
            "citations": [fabricated],
        },
    )
    assert rejected.json()["grounded"] is False
    assert rejected.json()["citation_integrity"][0]["verified"] is False


def test_verify_evidence_tool_is_tenant_scoped_and_cached(client: TestClient) -> None:
    project = client.post("/api/projects", json={"name": "Tenant A", "tenant_id": "tenant-a"}).json()
    project_id = project["id"]
    citation = {
        "document": "policy.txt",
        "page_or_sheet": "Line 1",
        "quote": "Backups are retained for 30 days.",
    }
    payload = {
        "tenant_id": "tenant-a",
        "project_id": project_id,
        "question": "How long are backups retained?",
        "answer": "Backups are retained for 30 days.",
        "citations": [citation],
    }
    first = client.post("/api/tools/verify-evidence", json=payload)
    second = client.post("/api/tools/verify-evidence", json=payload)
    assert first.status_code == 200
    assert first.json()["cached"] is False
    assert second.json()["cached"] is True
    assert first.json()["request_id"] != second.json()["request_id"]

    wrong_tenant = client.post("/api/tools/verify-evidence", json={**payload, "tenant_id": "tenant-b"})
    assert wrong_tenant.status_code == 404


def test_verify_evidence_requires_stored_project_scope(client: TestClient) -> None:
    response = client.post(
        "/api/tools/verify-evidence",
        json={
            "tenant_id": "demo",
            "question": "What is documented?",
            "answer": "A fabricated answer.",
            "citations": [{"document": "inline.txt", "page_or_sheet": "Line 1", "quote": "A fabricated answer."}],
        },
    )
    assert response.status_code == 422


def test_new_evidence_invalidates_prior_approval(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("current.md", b"Customers are notified within 48 hours after a confirmed incident.", "text/markdown"))],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["When are customers notified after a confirmed incident?"]},
    )
    question = client.post(f"/api/projects/{project_id}/run").json()["questions"][0]
    approved = client.patch(
        f"/api/projects/{project_id}/questions/{question['id']}/review",
        json={"action": "approve"},
    )
    assert approved.status_code == 200
    assert client.get(f"/api/projects/{project_id}/export?format=json").json()

    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("legacy.md", b"Customers are notified within 72 hours after a confirmed incident.", "text/markdown"))],
    )
    refreshed = client.get(f"/api/projects/{project_id}/questions").json()[0]
    assert refreshed["status"] == "draft"
    assert client.get(f"/api/projects/{project_id}/export?format=json").json() == []


def test_exports_escape_formula_like_values(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[("files", ("formula.txt", b"The documented formula is =1+1.", "text/plain"))],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["What is the documented formula?"]},
    )
    question = client.post(f"/api/projects/{project_id}/run").json()["questions"][0]
    approved = client.patch(
        f"/api/projects/{project_id}/questions/{question['id']}/review",
        json={"action": "approve", "edited_answer": "=1+1"},
    )
    assert approved.status_code == 200

    workbook = load_workbook(io.BytesIO(client.get(f"/api/projects/{project_id}/export?format=xlsx").content), data_only=False)
    assert workbook.active["B2"].value == "'=1+1"
    csv_text = client.get(f"/api/projects/{project_id}/export?format=csv").content.decode("utf-8-sig")
    assert "'=1+1" in csv_text


def test_questionnaire_upload_extracts_questions(client: TestClient) -> None:
    project_id = _project(client)
    response = client.post(
        f"/api/projects/{project_id}/documents?kind=questionnaire",
        files=[
            (
                "files",
                ("rfp.md", b"Do you support SSO?\nDescribe your incident response process.\n", "text/markdown"),
            )
        ],
    )
    assert response.status_code == 201
    assert len(response.json()["uploads"][0]["questions_added"]) == 2


def test_numeric_timeline_conflict_is_flagged(client: TestClient) -> None:
    project_id = _project(client)
    client.post(
        f"/api/projects/{project_id}/documents?kind=evidence",
        files=[
            ("files", ("current.md", b"Customers are notified within 48 hours after a confirmed incident.", "text/markdown")),
            ("files", ("legacy.md", b"Customers are notified within 72 hours after a confirmed incident.", "text/markdown")),
        ],
    )
    client.post(
        f"/api/projects/{project_id}/questions",
        json={"questions": ["When are customers notified after a confirmed incident?"]},
    )

    question = client.post(f"/api/projects/{project_id}/run").json()["questions"][0]

    assert question["checks"]["contradictions"]
    blocked = client.patch(
        f"/api/projects/{project_id}/questions/{question['id']}/review",
        json={"action": "approve"},
    )
    assert blocked.status_code == 422
    approved = client.patch(
        f"/api/projects/{project_id}/questions/{question['id']}/review",
        json={"action": "approve", "note": "Current 48-hour notice approved; legacy source is superseded."},
    )
    assert approved.status_code == 200
