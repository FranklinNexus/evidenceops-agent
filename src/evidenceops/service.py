from __future__ import annotations

import csv
import hashlib
import io
import json
import threading
import time
import uuid
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import Settings
from .models import (
    Citation,
    CitationIntegrity,
    DocumentKind,
    DocumentRecord,
    QuestionRecord,
    ReviewRequest,
    ReviewStatus,
    RunResponse,
    SaveDraftRequest,
    VerifyEvidenceRequest,
    VerifyEvidenceResponse,
)
from .parsers import extract_questions, parse_document, split_question_text
from .providers import AnswerProvider, DeterministicEvidenceProvider
from .retrieval import analyze_grounding, retrieve
from .store import NotFoundError, SQLiteStore


@dataclass(slots=True)
class UploadResult:
    document: DocumentRecord
    questions_added: list[QuestionRecord]


class TenantRateLimitExceeded(RuntimeError):
    pass


class EvidenceOpsService:
    """Reusable application boundary shared by HTTP, tests, and optional MCP transports."""

    def __init__(self, store: SQLiteStore, provider: AnswerProvider, settings: Settings) -> None:
        self.store = store
        self.provider = provider
        self.settings = settings
        self.grounded_fallback = DeterministicEvidenceProvider()
        self._tool_lock = threading.Lock()
        self._tool_events: dict[str, list[float]] = {}
        self._tool_cache: dict[str, tuple[float, VerifyEvidenceResponse]] = {}

    def upload_document(
        self,
        project_id: str,
        *,
        filename: str,
        data: bytes,
        kind: DocumentKind,
        content_type: str | None,
    ) -> UploadResult:
        if len(data) > self.settings.max_upload_mb * 1024 * 1024:
            raise ValueError(f"File exceeds {self.settings.max_upload_mb} MB upload limit")
        parsed = parse_document(filename, data)
        document = self.store.add_document(
            project_id,
            filename,
            kind,
            content_type,
            [chunk.model_dump() for chunk in parsed],
        )
        added: list[QuestionRecord] = []
        if kind == DocumentKind.questionnaire:
            extracted = extract_questions(parsed)
            added = self.store.add_questions(
                project_id,
                [question for question, _ in extracted],
                source_document=filename,
                source_locators=[locator for _, locator in extracted],
            )
        return UploadResult(document=document, questions_added=added)

    def add_manual_questions(self, project_id: str, questions: list[str], text: str | None) -> list[QuestionRecord]:
        combined = [question.strip() for question in questions if question.strip()]
        if text:
            combined.extend(split_question_text(text))
        return self.store.add_questions(project_id, combined)

    def run_project(self, project_id: str) -> RunResponse:
        chunks = self.store.list_chunks(project_id)
        questions = self.store.list_questions(project_id)
        processed = 0
        for question in questions:
            if question.status in {ReviewStatus.approved, ReviewStatus.rejected}:
                continue
            citations = retrieve(question.question, chunks, self.settings.retrieval_top_k)
            if not citations:
                self.store.update_question(
                    project_id,
                    question.id,
                    status=ReviewStatus.needs_evidence.value,
                    answer=None,
                    citations_json="[]",
                    checks_json=analyze_grounding(None, []).model_dump_json(),
                    missing_evidence=f"Evidence needed to answer: {question.question}",
                    provider=self.provider.name,
                )
                processed += 1
                continue
            draft = self.provider.draft(question.question, citations)
            selected = [citations[index] for index in draft.citation_indexes if 0 <= index < len(citations)]
            if not selected and draft.answer:
                selected = citations
            checks = analyze_grounding(draft.answer, selected)
            if draft.answer and checks.hallucination_risk:
                draft = self.grounded_fallback.draft(question.question, citations)
                selected = [citations[index] for index in draft.citation_indexes]
                checks = analyze_grounding(draft.answer, selected)
            status = ReviewStatus.draft if draft.answer and checks.grounded else ReviewStatus.needs_evidence
            missing = None if status == ReviewStatus.draft else f"More direct evidence is needed for: {question.question}"
            self.store.update_question(
                project_id,
                question.id,
                status=status.value,
                answer=draft.answer,
                citations_json=json.dumps([citation.model_dump() for citation in selected], ensure_ascii=False),
                checks_json=checks.model_dump_json(),
                missing_evidence=missing,
                provider=draft.provider,
            )
            processed += 1
        current = self.store.list_questions(project_id)
        return RunResponse(
            processed=processed,
            draft_count=sum(question.status == ReviewStatus.draft for question in current),
            needs_evidence_count=sum(question.status == ReviewStatus.needs_evidence for question in current),
            questions=current,
        )

    def review_question(self, project_id: str, question_id: str, review: ReviewRequest) -> QuestionRecord:
        question = self.store.get_question(project_id, question_id)
        answer = review.edited_answer.strip() if review.edited_answer else question.answer
        if review.action == "approve":
            if not answer:
                raise ValueError("An answer is required before approval")
            checks = analyze_grounding(answer, question.citations)
            if not checks.grounded:
                raise ValueError("Edited answer contains claims not supported by its citations")
            if checks.contradictions and not (review.note and review.note.strip()):
                raise ValueError("Conflicting evidence requires a reviewer disposition note before approval")
            status = ReviewStatus.approved
            checks_json = checks.model_dump_json()
        else:
            status = ReviewStatus.rejected
            checks_json = question.checks.model_dump_json() if question.checks else None
        return self.store.update_question(
            project_id,
            question_id,
            status=status.value,
            answer=answer,
            checks_json=checks_json,
            reviewer_note=review.note,
        )

    def save_draft(self, project_id: str, question_id: str, draft: SaveDraftRequest) -> QuestionRecord:
        question = self.store.get_question(project_id, question_id)
        answer = draft.edited_answer if draft.edited_answer is not None else question.answer
        checks = analyze_grounding(answer, question.citations)
        status = question.status
        if status == ReviewStatus.approved and (answer != question.answer or draft.note != question.reviewer_note):
            status = ReviewStatus.draft
        return self.store.update_question(
            project_id,
            question_id,
            status=status.value,
            answer=answer,
            checks_json=checks.model_dump_json(),
            reviewer_note=draft.note,
        )

    def verify_evidence(self, request: VerifyEvidenceRequest) -> VerifyEvidenceResponse:
        request_id = request.request_id or uuid.uuid4().hex
        request_payload = request.model_dump(exclude={"request_id"}, mode="json")
        request_hash = hashlib.sha256(
            json.dumps(request_payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        now = time.monotonic()
        with self._tool_lock:
            events = [event for event in self._tool_events.get(request.tenant_id, []) if now - event < 60]
            if len(events) >= self.settings.tool_calls_per_minute:
                self.store.record_tool_audit(
                    tenant_id=request.tenant_id,
                    request_id=request_id,
                    action="verify_evidence",
                    request_hash=request_hash,
                    cached=False,
                    outcome="rate_limited",
                )
                raise TenantRateLimitExceeded("Tenant verification call limit reached")
            events.append(now)
            self._tool_events[request.tenant_id] = events
            cached = self._tool_cache.get(request_hash)
            if cached and cached[0] > now:
                response = cached[1].model_copy(update={"request_id": request_id, "cached": True})
                self.store.record_tool_audit(
                    tenant_id=request.tenant_id,
                    request_id=request_id,
                    action="verify_evidence",
                    request_hash=request_hash,
                    cached=True,
                    outcome="complete",
                )
                return response

        if request.project_id and not self.store.project_belongs_to(request.project_id, request.tenant_id):
            self.store.record_tool_audit(
                tenant_id=request.tenant_id,
                request_id=request_id,
                action="verify_evidence",
                request_hash=request_hash,
                cached=False,
                outcome="project_scope_rejected",
            )
            raise NotFoundError("Project not found for tenant")

        citations = [Citation(**citation.model_dump()) for citation in request.citations]
        integrity: list[CitationIntegrity] = []
        for index, citation in enumerate(citations):
            if request.project_id:
                verified = self.store.find_quote(request.project_id, citation)
                reason = "Quote and locator match stored evidence" if verified else "Quote or locator not found in project evidence"
            else:
                verified = True
                reason = "Citation supplied inline; storage integrity was not requested"
            integrity.append(CitationIntegrity(citation_index=index, verified=verified, reason=reason))
        checks = analyze_grounding(request.answer, citations)
        all_integral = all(item.verified for item in integrity)
        response = VerifyEvidenceResponse(
            tenant_id=request.tenant_id,
            request_id=request_id,
            grounded=checks.grounded and all_integral,
            hallucination_risk=checks.hallucination_risk or not all_integral,
            unsupported_claims=checks.unsupported_claims,
            contradictions=checks.contradictions,
            citation_integrity=integrity,
        )
        with self._tool_lock:
            self._tool_cache[request_hash] = (now + self.settings.tool_cache_ttl_seconds, response)
        self.store.record_tool_audit(
            tenant_id=request.tenant_id,
            request_id=request_id,
            action="verify_evidence",
            request_hash=request_hash,
            cached=False,
            outcome="complete",
        )
        return response

    def export_project(
        self, project_id: str, export_format: str, *, include_drafts: bool = False
    ) -> tuple[bytes, str, str]:
        questions = self.store.list_questions(project_id)
        if not include_drafts:
            questions = [question for question in questions if question.status == ReviewStatus.approved]
        rows = [
            {
                "question": question.question,
                "answer": question.answer or "",
                "status": question.status.value,
                "citations": [
                    {
                        "document": citation.document,
                        "page_or_sheet": citation.page_or_sheet,
                        "quote": citation.quote,
                    }
                    for citation in question.citations
                ],
                "missing_evidence": question.missing_evidence or "",
                "reviewer_note": question.reviewer_note or "",
            }
            for question in questions
        ]
        if export_format == "json":
            body = json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")
            return body, "application/json", "evidenceops-export.json"
        if export_format == "xlsx":
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Approved answers" if not include_drafts else "Questionnaire answers"
            fieldnames = ["question", "answer", "status", "citations", "missing_evidence", "reviewer_note"]
            worksheet.append(fieldnames)
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="17324D")
            for row in rows:
                worksheet.append(
                    [
                        json.dumps(row[field], ensure_ascii=False) if field == "citations" else row[field]
                        for field in fieldnames
                    ]
                )
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            widths = {"A": 48, "B": 64, "C": 16, "D": 64, "E": 42, "F": 36}
            for column, width in widths.items():
                worksheet.column_dimensions[column].width = width
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            output = io.BytesIO()
            workbook.save(output)
            return (
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "evidenceops-export.xlsx",
            )
        output = io.StringIO(newline="")
        fieldnames = ["question", "answer", "status", "citations", "missing_evidence", "reviewer_note"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(
            {**row, "citations": json.dumps(row["citations"], ensure_ascii=False)}
            for row in rows
        )
        return output.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", "evidenceops-export.csv"
