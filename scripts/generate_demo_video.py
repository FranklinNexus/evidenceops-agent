from __future__ import annotations

import asyncio
import json
import math
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path

import edge_tts
import imageio_ffmpeg
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
CAPTURES = ROOT / "work" / "video-captures"
BUILD = ROOT / "work" / "demo-video"
OUTPUT = ROOT / "outputs" / "evidenceops-demo-video.mp4"
SUBTITLES = ROOT / "outputs" / "evidenceops-demo-video.en.srt"
EXPORT = ROOT / "outputs" / "evidenceops-demo-export.xlsx"
WIDTH = 1920
HEIGHT = 1080
FPS = 30
VOICE = "en-US-AriaNeural"
TRANSITION = 0.35

INK = "#182321"
PAPER = "#F4F7F5"
WHITE = "#FFFFFF"
TEAL = "#08766B"
TEAL_SOFT = "#DDF1EC"
CORAL = "#B14B4B"
CORAL_SOFT = "#F8E5E2"
AMBER = "#9B6500"
AMBER_SOFT = "#F7E9C9"
BLUE = "#355A77"
BLUE_SOFT = "#E5EEF5"
MUTED = "#64706D"


@dataclass(frozen=True)
class Scene:
    slug: str
    title: str
    kicker: str
    narration: str
    kind: str
    asset: str | None = None
    highlights: tuple[tuple[int, int, int, int, str], ...] = ()


SCENES = [
    Scene(
        "00-title",
        "EvidenceOps",
        "THE AGENT THAT KNOWS WHEN NOT TO ANSWER",
        "Meet EvidenceOps: the compliance questionnaire agent that knows when not to answer. "
        "Security reviews are not primarily a writing problem; they are an evidence-control problem. "
        "A fluent answer without proof can create audit and contract risk. EvidenceOps makes citations, "
        "uncertainty, and human approval part of one workflow.",
        "title",
    ),
    Scene(
        "01-boundary",
        "A deliberate trust boundary",
        "SYNTHETIC DATA / LOCAL MVP",
        "This is a runnable local MVP using a fully synthetic CloudDesk dataset; no customer data or secrets "
        "appear anywhere in this video. A reviewer loads a questionnaire and an organization-controlled evidence "
        "library. The interface accepts PDF, Excel, CSV, Word, Markdown, and text. Uploaded documents are evidence. "
        "Generated text is not.",
        "ui",
        "00-empty-workspace.png",
        ((14, 94, 426, 1064, TEAL),),
    ),
    Scene(
        "02-pipeline",
        "One run, four controlled stages",
        "EXTRACT / RETRIEVE / DRAFT",
        "One run extracts eight traceable questions, retrieves from the four indexed documents, drafts only from "
        "retrieved passages, verifies claims, and classifies gaps. This recording uses the local evidence-only "
        "provider. The same server-side provider boundary can invoke Strands with an operator-controlled "
        "OpenAI-compatible endpoint or Bedrock, but no hosted model or AWS deployment is being claimed here.",
        "ui",
        "00-agent-running.png",
        ((16, 742, 422, 1068, TEAL),),
    ),
    Scene(
        "03-citation",
        "A citation is a control, not decoration",
        "ANSWER + EXACT SOURCE PASSAGE",
        "Open the encryption question. The draft states TLS 1.2 or later in transit and AES-256 at rest. Directly "
        "below it, the citation resolves to security-overview dot M D, with the exact quoted passage and stable "
        "locator. The answer is human-approved, but the source remains visible so a reviewer can verify the claim "
        "instead of trusting model confidence.",
        "ui",
        "01-workspace.png",
        ((1050, 225, 1810, 535, TEAL), (1050, 748, 1810, 902, TEAL)),
    ),
    Scene(
        "04-conflict",
        "Conflict stays visible",
        "48 HOURS VS 72 HOURS",
        "More important than a polished answer is a visible disagreement. The current privacy addendum promises "
        "notice no later than forty-eight hours, while a superseded incident plan says seventy-two. EvidenceOps "
        "retains both citations, labels the conflict, and requires a reviewer disposition note before approval. "
        "It does not silently pick the more convenient answer.",
        "ui",
        "02-conflict.png",
        ((1050, 480, 1812, 952, CORAL),),
    ),
    Scene(
        "05-gap",
        "Missing proof becomes explicit work",
        "0 WORDS / 0 CITATIONS",
        "There is no current subprocessor register in this evidence pack. EvidenceOps leaves the answer empty, "
        "shows zero citations, and creates a specific evidence request tied to the question. The item stays in "
        "needs review and cannot be approved as final. Missing proof remains missing instead of becoming plausible prose.",
        "ui",
        "03-missing-evidence.png",
        ((1050, 724, 1812, 928, AMBER),),
    ),
    Scene(
        "06-approval",
        "A person makes the decision",
        "SUPPORTED DRAFT -> APPROVED",
        "For a supported multi-factor authentication draft, the reviewer can compare the wording with the exact "
        "security overview quote and explicitly approve it. Approval is a persisted workflow event, not a decorative "
        "button. By default, only approved responses are eligible for final export.",
        "ui",
        "05-approved.png",
        ((1136, 112, 1240, 165, TEAL), (1050, 225, 1810, 535, TEAL), (1050, 748, 1810, 902, TEAL), (1668, 991, 1852, 1068, TEAL)),
    ),
    Scene(
        "07-invalidation",
        "Edits invalidate stale sign-off",
        "APPROVAL CLEARED AUTOMATICALLY",
        "Now I edit the approved wording. The interface immediately returns the item to Draft, and the approved "
        "count drops. The old sign-off is invalidated instead of following changed text. The citation remains attached, "
        "the edited claim is checked again, and the final evidence-backed wording requires a new approval.",
        "ui",
        "06-approval-invalidated.png",
        ((1136, 112, 1240, 165, CORAL), (1050, 225, 1810, 535, CORAL), (1050, 748, 1810, 902, CORAL), (1668, 991, 1852, 1068, TEAL)),
    ),
    Scene(
        "08-export",
        "Structured export, never auto-submission",
        "XLSX / CSV / JSON",
        "The export dialog offers an Excel workbook, portable C S V, or structured JSON. Approved-only is the default. For this demo, "
        "including drafts lets the workbook carry each question, answer, decision status, citations, missing evidence, "
        "and reviewer note. EvidenceOps prepares the handoff; it never auto-submits a response to a buyer.",
        "ui",
        "08-export-modal.png",
        ((620, 230, 1300, 850, TEAL),),
    ),
    Scene(
        "09-workbook",
        "The exported workbook preserves context",
        "ANSWERS / STATUS / CITATIONS / GAPS",
        "The generated workbook keeps approved answers beside their exact source excerpts, while unresolved work "
        "retains its draft or needs-evidence status. The subprocessor row carries the evidence request instead of a "
        "fabricated list. That makes the output useful to reviewers without presenting unfinished work as fact.",
        "workbook",
    ),
    Scene(
        "10-architecture",
        "Agentic drafting, deterministic controls",
        "STRANDS INSIDE A HUMAN-GATED SERVICE",
        "The architecture intentionally divides generative and deterministic responsibilities. Strands owns the "
        "model-assisted draft step through a typed AgentDraft result: answer text plus selected citation indexes. "
        "Deterministic code owns document identity, retrieval, grounding, numeric conflict checks, gaps, approval "
        "transitions, and export. The provider wrapper adds tenant limits, bounded retries, a circuit breaker, and an "
        "evidence-only fallback. An outage can produce a grounded quote or an explicit gap, never an uncited guess.",
        "architecture",
    ),
    Scene(
        "11-strands",
        "The Strands path is release-gated",
        "REAL SDK / LOOPBACK PROTOCOL FIXTURE",
        "The release gate uses the installed Strands SDK, constructs a Strands OpenAIModel and Agent, sends the typed "
        "tool schema to a loopback OpenAI-compatible fixture, receives selected citation indexes, then runs retrieval "
        "and grounding through the same EvidenceOps service. The test passed. This proves the SDK integration path; "
        "it is not a claim of hosted model quality, Bedrock use, or AgentCore deployment.",
        "validation",
    ),
    Scene(
        "12-close",
        "No evidence, no answer.",
        "LOCAL MVP TODAY / CLEAR AWS PATH NEXT",
        "The working local stack is FastAPI, SQLite, a responsive review workspace, and the Strands provider boundary. "
        "The architecture maps an optional future path to S3, DynamoDB, CloudWatch, managed retrieval, and AgentCore or "
        "a container service; those are deployment options, not live claims. EvidenceOps delivers an honest, cited, "
        "human-approved workbook.",
        "close",
    ),
]


def font(size: int, *, bold: bool = False, mono: bool = False) -> ImageFont.FreeTypeFont:
    if mono:
        path = Path("C:/Windows/Fonts/consola.ttf")
    elif bold:
        path = Path("C:/Windows/Fonts/seguisb.ttf")
    else:
        path = Path("C:/Windows/Fonts/segoeui.ttf")
    return ImageFont.truetype(str(path), size)


def rounded(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int], radius: int, **kwargs: object) -> None:
    draw.rounded_rectangle(box, radius=radius, **kwargs)


def wrap(draw: ImageDraw.ImageDraw, text: str, text_font: ImageFont.FreeTypeFont, max_width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if draw.textbbox((0, 0), candidate, font=text_font)[2] <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def draw_text_block(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    text_font: ImageFont.FreeTypeFont,
    fill: str,
    max_width: int,
    spacing: int = 10,
) -> int:
    x, y = xy
    line_height = text_font.size + spacing
    for line in wrap(draw, text, text_font, max_width):
        draw.text((x, y), line, font=text_font, fill=fill)
        y += line_height
    return y


def add_scene_badges(image: Image.Image, scene: Scene) -> None:
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    title_font = font(25, bold=True)
    kicker_font = font(17, bold=True)
    title_box = (1130, 12, 1686, 61)
    kicker_box = (840, 15, 1115, 58)
    rounded(draw, title_box, 9, fill=(24, 35, 33, 236))
    rounded(draw, kicker_box, 9, fill=(221, 241, 236, 244), outline=(8, 118, 107, 150), width=2)
    draw.text((1151, 21), scene.title, font=title_font, fill=WHITE)
    kicker = scene.kicker[:31]
    bbox = draw.textbbox((0, 0), kicker, font=kicker_font)
    draw.text((977 - (bbox[2] - bbox[0]) // 2, 26), kicker, font=kicker_font, fill=TEAL)
    image.alpha_composite(layer)


def add_highlights(image: Image.Image, highlights: tuple[tuple[int, int, int, int, str], ...]) -> None:
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    for x1, y1, x2, y2, color in highlights:
        rgb = ImageColor_getrgb(color)
        rounded(draw, (x1, y1, x2, y2), 8, fill=(*rgb, 18), outline=(*rgb, 235), width=5)
    image.alpha_composite(layer)


def ImageColor_getrgb(value: str) -> tuple[int, int, int]:
    value = value.lstrip("#")
    return tuple(int(value[index : index + 2], 16) for index in (0, 2, 4))  # type: ignore[return-value]


def ui_slide(scene: Scene) -> Image.Image:
    source = Image.open(CAPTURES / str(scene.asset)).convert("RGBA")
    if source.size != (WIDTH, HEIGHT):
        source = source.resize((WIDTH, HEIGHT), Image.Resampling.LANCZOS)
    add_highlights(source, scene.highlights)
    add_scene_badges(source, scene)
    return source.convert("RGB")


def title_slide(scene: Scene, *, close: bool = False) -> Image.Image:
    image = Image.new("RGB", (WIDTH, HEIGHT), INK)
    draw = ImageDraw.Draw(image)
    # Restrained geometric accents echo evidence, review, and exception states.
    rounded(draw, (86, 84, 196, 194), 22, fill=TEAL)
    draw.rectangle((122, 111, 159, 147), outline=WHITE, width=4)
    draw.line((116, 163, 140, 178, 170, 139), fill=WHITE, width=6, joint="curve")
    draw.text((236, 86), "EvidenceOps", font=font(76, bold=True), fill=WHITE)
    draw.text((240, 177), "Questionnaire review", font=font(28), fill="#B9C7C3")
    draw.rectangle((86, 266, 1840, 268), fill="#32413E")
    headline = scene.title if close else "The agent that knows when not to answer."
    draw_text_block(draw, (86, 332), headline, font(62, bold=True), WHITE, 1450, spacing=16)
    draw_text_block(
        draw,
        (90, 500),
        "Cited drafts. Visible conflicts. Explicit evidence gaps. Human approval before export.",
        font(34),
        "#D5DFDC",
        1420,
        spacing=14,
    )
    badges = [("CITATIONS", TEAL_SOFT, TEAL), ("CONFLICTS", CORAL_SOFT, CORAL), ("HUMAN GATE", AMBER_SOFT, AMBER)]
    x = 90
    for label, bg, fg in badges:
        rounded(draw, (x, 668, x + 250, 726), 11, fill=bg)
        bbox = draw.textbbox((0, 0), label, font=font(21, bold=True))
        draw.text((x + 125 - (bbox[2] - bbox[0]) / 2, 683), label, font=font(21, bold=True), fill=fg)
        x += 272
    rounded(draw, (86, 884, 1838, 982), 14, fill="#22302D", outline="#3B4D49", width=2)
    draw.text((118, 910), "LOCAL MVP", font=font(21, bold=True), fill="#6FD3C4")
    draw.text((300, 910), "Synthetic fixtures", font=font(21), fill=WHITE)
    draw.text((582, 910), "No customer data", font=font(21), fill=WHITE)
    draw.text((860, 910), "No live AWS or hosted-model claim", font=font(21), fill=WHITE)
    draw.text((90, 1020), "github.com/FranklinNexus/evidenceops-agent", font=font(22, mono=True), fill="#A9BBB6")
    return image


def workbook_slide(scene: Scene) -> Image.Image:
    workbook = load_workbook(EXPORT, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    selected = [rows[1], rows[4], rows[-1]]
    image = Image.new("RGB", (WIDTH, HEIGHT), PAPER)
    draw = ImageDraw.Draw(image)
    draw.text((72, 55), scene.title, font=font(46, bold=True), fill=INK)
    draw.text((74, 116), "A real XLSX generated by the local API", font=font(23), fill=MUTED)
    rounded(draw, (72, 180, 1848, 930), 15, fill=WHITE, outline="#CCD7D3", width=2)
    headers = ["Question", "Answer / evidence state", "Status", "Citation or missing evidence"]
    widths = [440, 620, 185, 485]
    starts = [96, 536, 1156, 1341]
    draw.rectangle((73, 181, 1847, 251), fill="#E8EFEC")
    for x, label in zip(starts, headers):
        draw.text((x, 202), label, font=font(20, bold=True), fill=INK)
    row_y = 252
    heights = [210, 210, 255]
    for index, (row, row_height) in enumerate(zip(selected, heights)):
        question, answer, status, citations, missing, _note = row
        fill = "#FBFCFB" if index % 2 == 0 else "#F2F6F4"
        draw.rectangle((73, row_y, 1847, row_y + row_height), fill=fill)
        for x in (516, 1136, 1321):
            draw.line((x, row_y, x, row_y + row_height), fill="#D8E0DD", width=2)
        question_lines = wrap(draw, str(question), font(21, bold=True), widths[0] - 35)[:4]
        answer_lines = wrap(draw, str(answer or "No answer drafted"), font(19), widths[1] - 36)[:5]
        y = row_y + 27
        for line in question_lines:
            draw.text((96, y), line, font=font(21, bold=True), fill=INK)
            y += 30
        y = row_y + 27
        for line in answer_lines:
            draw.text((536, y), line, font=font(19), fill="#34413E")
            y += 29
        status_text = str(status).replace("_", " ").upper()
        if status == "approved":
            status_bg, status_fg = TEAL_SOFT, TEAL
        elif status == "needs_evidence":
            status_bg, status_fg = AMBER_SOFT, AMBER
        else:
            status_bg, status_fg = BLUE_SOFT, BLUE
        rounded(draw, (1160, row_y + 27, 1304, row_y + 73), 9, fill=status_bg)
        status_bbox = draw.textbbox((0, 0), status_text, font=font(16, bold=True))
        draw.text((1232 - (status_bbox[2] - status_bbox[0]) / 2, row_y + 40), status_text, font=font(16, bold=True), fill=status_fg)
        context = str(missing or "")
        if not context and citations:
            try:
                first = json.loads(str(citations))[0]
                context = f"{first['document']} - {first['page_or_sheet']}\n{first['quote']}"
            except (json.JSONDecodeError, KeyError, IndexError, TypeError):
                context = str(citations)
        context_lines = wrap(draw, context, font(17), widths[3] - 38)[:7]
        y = row_y + 27
        for line in context_lines:
            draw.text((1341, y), line, font=font(17), fill="#44514E")
            y += 27
        row_y += row_height
    rounded(draw, (72, 968, 1848, 1032), 11, fill=INK)
    draw.text((100, 986), "Approved-only is the default. Drafts are included here only to demonstrate review context.", font=font(22), fill=WHITE)
    return image


def arrow(draw: ImageDraw.ImageDraw, start: tuple[int, int], end: tuple[int, int], color: str = MUTED) -> None:
    draw.line((*start, *end), fill=color, width=4)
    angle = math.atan2(end[1] - start[1], end[0] - start[0])
    for offset in (2.55, -2.55):
        point = (end[0] + 15 * math.cos(angle + offset), end[1] + 15 * math.sin(angle + offset))
        draw.line((*end, *point), fill=color, width=4)


def architecture_slide(scene: Scene) -> Image.Image:
    image = Image.new("RGB", (WIDTH, HEIGHT), PAPER)
    draw = ImageDraw.Draw(image)
    draw.text((72, 48), scene.title, font=font(44, bold=True), fill=INK)
    draw.text((74, 105), "The model proposes. Deterministic controls and a person decide.", font=font(23), fill=MUTED)
    labels = [
        ("1", "Questionnaire\n+ evidence", AMBER_SOFT, AMBER),
        ("2", "Parse +\nretrieve", BLUE_SOFT, BLUE),
        ("3", "Strands\nAgentDraft", TEAL_SOFT, TEAL),
        ("4", "Grounding +\nconflicts", BLUE_SOFT, BLUE),
        ("5", "Human\nreview", AMBER_SOFT, AMBER),
        ("6", "Approved\nexport", TEAL_SOFT, TEAL),
    ]
    x_positions = [76, 372, 668, 964, 1260, 1556]
    box_y = 236
    for index, ((number, label, bg, fg), x) in enumerate(zip(labels, x_positions)):
        rounded(draw, (x, box_y, x + 245, box_y + 170), 16, fill=WHITE, outline="#C7D3CF", width=2)
        rounded(draw, (x + 18, box_y + 18, x + 62, box_y + 62), 9, fill=bg)
        draw.text((x + 33, box_y + 27), number, font=font(18, bold=True), fill=fg, anchor="mm")
        lines = label.split("\n")
        for line_index, line in enumerate(lines):
            draw.text((x + 122, box_y + 88 + line_index * 32), line, font=font(24, bold=True), fill=INK, anchor="mm")
        if index < len(labels) - 1:
            arrow(draw, (x + 247, box_y + 85), (x_positions[index + 1] - 4, box_y + 85))
    rounded(draw, (344, 500, 930, 815), 16, fill=WHITE, outline="#B9CFC8", width=2)
    draw.text((374, 531), "Server-side provider boundary", font=font(28, bold=True), fill=INK)
    provider_items = [
        "Operator-controlled OpenAI-compatible adapter",
        "Optional Amazon Bedrock adapter",
        "Rate limit + retry + circuit breaker",
        "Evidence-only deterministic fallback",
    ]
    y = 593
    for item in provider_items:
        rounded(draw, (374, y, 397, y + 23), 5, fill=TEAL_SOFT, outline=TEAL)
        draw.text((417, y - 3), item, font=font(21), fill="#37433F")
        y += 52
    arrow(draw, (790, 500), (790, 413), TEAL)
    rounded(draw, (990, 500, 1577, 815), 16, fill=WHITE, outline="#C8D0D8", width=2)
    draw.text((1020, 531), "Deterministic application core", font=font(28, bold=True), fill=INK)
    controls = [
        "Stable document and citation identity",
        "Evidence-only retrieval and integrity checks",
        "Numeric conflict and unsupported-claim checks",
        "Approval invalidation after edits",
        "Approved-only export gate",
    ]
    y = 593
    for item in controls:
        rounded(draw, (1020, y, 1043, y + 23), 5, fill=BLUE_SOFT, outline=BLUE)
        draw.text((1063, y - 3), item, font=font(21), fill="#37433F")
        y += 45
    arrow(draw, (1285, 500), (1285, 413), BLUE)
    rounded(draw, (72, 882, 1848, 1026), 14, fill=INK)
    draw.text((104, 910), "IMPLEMENTED LOCAL MVP", font=font(20, bold=True), fill="#6FD3C4")
    draw.text((104, 948), "FastAPI  |  SQLite  |  Strands provider adapters  |  Web review workspace", font=font(24), fill=WHITE)
    draw.text((1050, 910), "OPTIONAL AWS PATH", font=font(20, bold=True), fill="#F1C76A")
    draw.text((1050, 948), "AgentCore / ECS  |  S3  |  DynamoDB  |  CloudWatch", font=font(24), fill=WHITE)
    return image


def validation_slide(scene: Scene) -> Image.Image:
    image = Image.new("RGB", (WIDTH, HEIGHT), "#16201F")
    draw = ImageDraw.Draw(image)
    draw.text((72, 54), scene.title, font=font(44, bold=True), fill=WHITE)
    draw.text((74, 112), "Installed Strands SDK exercised end to end through EvidenceOpsService", font=font(23), fill="#B8C8C3")
    rounded(draw, (72, 190, 1160, 880), 16, fill="#0C1312", outline="#3B4D48", width=2)
    draw.rectangle((72, 190, 1160, 252), fill="#24302E")
    for x, color in ((105, CORAL), (139, "#D7A23E"), (173, TEAL)):
        draw.ellipse((x - 9, 212, x + 9, 230), fill=color)
    draw.text((207, 206), "pytest - Strands release gate", font=font(20, mono=True), fill="#D7E1DE")
    terminal = [
        "$ .venv/Scripts/python -m pytest",
        "  tests/test_strands_integration.py -vv",
        "",
        "collected 1 item",
        "",
        "test_strands_openai_provider_runs_",
        "through_evidenceops_service PASSED [100%]",
        "",
        "1 passed, 1 warning in 2.69s",
    ]
    y = 294
    for line in terminal:
        color = "#73D6C5" if "PASSED" in line or line.startswith("1 passed") else "#D5E0DD"
        draw.text((108, y), line, font=font(25, mono=True), fill=color)
        y += 54
    rounded(draw, (1204, 190, 1848, 880), 16, fill=WHITE)
    draw.text((1240, 232), "What the gate proves", font=font(30, bold=True), fill=INK)
    bullets = [
        "Real Strands Agent and OpenAIModel construction",
        "Typed AgentDraft tool schema sent over HTTP",
        "Citation indexes returned to the application",
        "Retrieval and grounding run through the shared service",
    ]
    y = 314
    for number, item in enumerate(bullets, 1):
        rounded(draw, (1242, y, 1284, y + 42), 8, fill=TEAL_SOFT)
        draw.text((1263, y + 20), str(number), font=font(17, bold=True), fill=TEAL, anchor="mm")
        y = draw_text_block(draw, (1304, y - 1), item, font(21), INK, 500, spacing=8) + 35
    rounded(draw, (1238, 738, 1814, 837), 11, fill=CORAL_SOFT)
    draw.text((1266, 760), "Disclosure", font=font(18, bold=True), fill=CORAL)
    draw.text((1266, 794), "Loopback fixture - no hosted LLM, Bedrock, or AgentCore call", font=font(17), fill="#6F3636")
    draw.text((74, 948), "Reproduce: .venv/Scripts/python -m pytest tests/test_strands_integration.py -vv", font=font(22, mono=True), fill="#9CB1AB")
    return image


def make_slide(scene: Scene) -> Image.Image:
    if scene.kind == "ui":
        return ui_slide(scene)
    if scene.kind == "workbook":
        return workbook_slide(scene)
    if scene.kind == "architecture":
        return architecture_slide(scene)
    if scene.kind == "validation":
        return validation_slide(scene)
    return title_slide(scene, close=scene.kind == "close")


async def synthesize(scene: Scene, destination: Path) -> None:
    communicator = edge_tts.Communicate(scene.narration, VOICE, rate="+12%", volume="+0%", pitch="-2Hz")
    await communicator.save(str(destination))


def duration(ffmpeg: str, media: Path) -> float:
    result = subprocess.run([ffmpeg, "-i", str(media)], capture_output=True, text=True, encoding="utf-8", errors="replace")
    match = re.search(r"Duration: (\d+):(\d+):(\d+(?:\.\d+)?)", result.stderr)
    if not match:
        raise RuntimeError(f"Could not read media duration for {media}")
    hours, minutes, seconds = match.groups()
    return int(hours) * 3600 + int(minutes) * 60 + float(seconds)


def timestamp(value: float) -> str:
    milliseconds = int(round(value * 1000))
    hours, remainder = divmod(milliseconds, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    seconds, milliseconds = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d},{milliseconds:03d}"


def build_subtitles(durations: list[float]) -> None:
    entries: list[str] = []
    start = 0.0
    index = 1
    for scene, scene_duration in zip(SCENES, durations):
        sentences = [part.strip() for part in re.split(r"(?<=[.!?])\s+", scene.narration) if part.strip()]
        total_chars = max(1, sum(len(sentence) for sentence in sentences))
        scene_audio = scene_duration - 0.7
        cursor = start + 0.15
        for sentence in sentences:
            share = max(1.8, scene_audio * len(sentence) / total_chars)
            end = min(start + scene_audio, cursor + share)
            entries.append(f"{index}\n{timestamp(cursor)} --> {timestamp(end)}\n{sentence}\n")
            cursor = end
            index += 1
        start += scene_duration - TRANSITION
    SUBTITLES.write_text("\n".join(entries), encoding="utf-8")


def render_segment(ffmpeg: str, image: Path, audio: Path, target: Path, audio_duration: float, *, intro: bool, outro: bool) -> float:
    total = audio_duration + 0.7
    video_filters = [f"scale={WIDTH}:{HEIGHT}:flags=lanczos", "format=yuv420p"]
    if intro:
        video_filters.append("fade=t=in:st=0:d=0.45")
    if outro:
        video_filters.append(f"fade=t=out:st={max(0.0, total - 0.55):.3f}:d=0.55")
    audio_filters = ["aresample=48000", "apad=pad_dur=0.7", "afade=t=in:st=0:d=0.12"]
    if outro:
        audio_filters.append(f"afade=t=out:st={max(0.0, audio_duration - 0.3):.3f}:d=0.3")
    command = [
        ffmpeg,
        "-y",
        "-loop",
        "1",
        "-framerate",
        str(FPS),
        "-i",
        str(image),
        "-i",
        str(audio),
        "-filter_complex",
        f"[0:v]{','.join(video_filters)}[v];[1:a]{','.join(audio_filters)}[a]",
        "-map",
        "[v]",
        "-map",
        "[a]",
        "-t",
        f"{total:.3f}",
        "-r",
        str(FPS),
        "-c:v",
        "libx264",
        "-preset",
        "medium",
        "-crf",
        "19",
        "-c:a",
        "aac",
        "-b:a",
        "192k",
        "-ar",
        "48000",
        "-movflags",
        "+faststart",
        str(target),
    ]
    subprocess.run(command, check=True, capture_output=True)
    return total


def combine(ffmpeg: str, segments: list[Path], durations: list[float]) -> None:
    command = [ffmpeg, "-y"]
    for segment in segments:
        command.extend(["-i", str(segment)])
    filters: list[str] = []
    video_label = "0:v"
    audio_label = "0:a"
    combined_duration = durations[0]
    for index in range(1, len(segments)):
        offset = combined_duration - TRANSITION
        next_video = f"v{index}"
        next_audio = f"a{index}"
        filters.append(
            f"[{video_label}][{index}:v]xfade=transition=fade:duration={TRANSITION}:offset={offset:.3f}[{next_video}]"
        )
        filters.append(
            f"[{audio_label}][{index}:a]acrossfade=d={TRANSITION}:c1=tri:c2=tri[{next_audio}]"
        )
        video_label = next_video
        audio_label = next_audio
        combined_duration += durations[index] - TRANSITION
    command.extend(
        [
            "-filter_complex",
            ";".join(filters),
            "-map",
            f"[{video_label}]",
            "-map",
            f"[{audio_label}]",
            "-c:v",
            "libx264",
            "-preset",
            "medium",
            "-crf",
            "18",
            "-pix_fmt",
            "yuv420p",
            "-c:a",
            "aac",
            "-b:a",
            "192k",
            "-movflags",
            "+faststart",
            str(OUTPUT),
        ]
    )
    subprocess.run(command, check=True, capture_output=True)


def main() -> None:
    missing = [scene.asset for scene in SCENES if scene.asset and not (CAPTURES / scene.asset).is_file()]
    if missing:
        raise FileNotFoundError(f"Missing browser captures: {', '.join(str(item) for item in missing)}")
    if not EXPORT.is_file():
        raise FileNotFoundError(f"Missing demo export: {EXPORT}")
    BUILD.mkdir(parents=True, exist_ok=True)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    segments: list[Path] = []
    segment_durations: list[float] = []
    for index, scene in enumerate(SCENES):
        slide_path = BUILD / f"{scene.slug}.png"
        audio_path = BUILD / f"{scene.slug}.mp3"
        segment_path = BUILD / f"{scene.slug}.mp4"
        make_slide(scene).save(slide_path, optimize=True)
        asyncio.run(synthesize(scene, audio_path))
        audio_duration = duration(ffmpeg, audio_path)
        scene_duration = render_segment(
            ffmpeg,
            slide_path,
            audio_path,
            segment_path,
            audio_duration,
            intro=index == 0,
            outro=index == len(SCENES) - 1,
        )
        segments.append(segment_path)
        segment_durations.append(scene_duration)
        print(f"{scene.slug}: {scene_duration:.2f}s")
    combine(ffmpeg, segments, segment_durations)
    build_subtitles(segment_durations)
    final_duration = duration(ffmpeg, OUTPUT)
    if final_duration > 300:
        raise RuntimeError(f"Final video is {final_duration:.2f}s; Devpost limit is 300s")
    print(f"Created {OUTPUT} ({final_duration:.2f}s, {WIDTH}x{HEIGHT}, H.264/AAC)")
    print(f"Created {SUBTITLES}")


if __name__ == "__main__":
    main()
